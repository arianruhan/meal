#!/usr/bin/env python3
"""
Downloads the live workbook from OneDrive and reads it in its ORIGINAL
layout — the same merged-cell, formula-driven "Meal_Chart_*.xlsx" format
you already use. No re-formatting or extra tabs needed; keep filling out
the sheet exactly as before and this script reads whatever Excel already
computed.

Required environment variable:
  ONEDRIVE_SHARE_URL - the "Anyone with the link can view" share link from
                        OneDrive's Share dialog. Paste it as-is; this script
                        converts it to a direct-download URL itself.

Layout assumptions (matches the original file):
  - One sheet, containing a block of rows per person:
      row      -> Breakfast day counts (counts as half a meal), plus that
                  person's Name (col A), Personal Total, Personal Cost,
                  Deposit, Balance, Khalar Bill, Gas Bill, Electricity Bill
      row + 1  -> Lunch day counts (full meal)
      row + 2  -> Dinner day counts (full meal)
      row + 3  -> blank spacer row
    Day counts live in columns B through AF (day 1 through day 31).
  - A "Bazar List" section below the person blocks, with a header row
    (Date / Cost / Person / Main Items) followed by one row per trip.
  - A summary block with labelled cells (e.g. "Grand Total (Meal)",
    "Meal Rate", "Total Deposit", "Current Balance", "Total due", and the
    electricity breakdown) — found by searching for the label text rather
    than hardcoded coordinates, so it keeps working if rows shift.

If your sheet is ever restructured significantly, this script will need
updating to match — it mirrors today's file exactly.
"""
import base64
import json
import os
import sys
import tempfile
import urllib.request
from datetime import datetime, timezone

import openpyxl

ONEDRIVE_SHARE_URL = os.environ["ONEDRIVE_SHARE_URL"]
SHEET_NAME = os.environ.get("SHEET_NAME")  # optional; defaults to the first sheet

DAY_COL_START = 2   # column B = Day 1
DAY_COL_END = 32     # column AF = Day 31


def onedrive_direct_download_url(share_url: str) -> str:
    b64 = base64.urlsafe_b64encode(share_url.strip().encode("utf-8")).decode("utf-8")
    b64 = b64.rstrip("=")
    encoded = "u!" + b64
    return f"https://api.onedrive.com/v1.0/shares/{encoded}/root/content"


def download_workbook(share_url: str) -> str:
    url = onedrive_direct_download_url(share_url)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    tmp_path = os.path.join(tempfile.gettempdir(), "meal_chart_live.xlsx")
    with urllib.request.urlopen(req, timeout=60) as resp, open(tmp_path, "wb") as out:
        out.write(resp.read())
    return tmp_path


def to_float(v, default=0.0):
    try:
        if v is None or str(v).strip() in ("", "\u200b"):
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


def norm(s):
    """Lowercase + collapse whitespace, for tolerant label matching."""
    return " ".join(str(s).split()).strip().lower() if s is not None else ""


def find_label_value(ws, label, max_cols_right=6, occurrence=0):
    """Search the whole sheet for a cell whose text matches `label`
    (case/space tolerant), then return the first non-empty cell up to
    `max_cols_right` columns to its right in the same row."""
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) == norm(label):
                matches.append(cell)
    if not matches or occurrence >= len(matches):
        return None
    label_cell = matches[occurrence]
    for offset in range(1, max_cols_right + 1):
        v = ws.cell(row=label_cell.row, column=label_cell.column + offset).value
        if v is not None:
            return v
    return None


def parse_people(ws, bazar_start_row):
    people = []
    r = 2
    while r < bazar_start_row:
        name_cell = ws.cell(row=r, column=1).value
        name = str(name_cell).strip() if name_cell is not None else ""
        if not name or name.upper() == "N/A":
            r += 4
            continue

        def day_counts(row):
            return {
                str(c - DAY_COL_START + 1): to_float(ws.cell(row=row, column=c).value)
                for c in range(DAY_COL_START, DAY_COL_END + 1)
            }

        breakfast = day_counts(r)
        lunch = day_counts(r + 1)
        dinner = day_counts(r + 2)

        days = {}
        for d in breakfast.keys():
            b, l, din = breakfast[d], lunch.get(d, 0.0), dinner.get(d, 0.0)
            days[d] = {"breakfast": b, "lunch": l, "dinner": din, "total": b * 0.5 + l + din}

        # Personal Total / Cost / Deposit / Balance / Khalar / Gas / Electricity
        # live on the same (first) row of the block. Column letters can drift
        # if the sheet changes, so find them by header label once per file
        # rather than hardcoding — see header_cols below.
        people.append({
            "name": name,
            "row": r,
            "days": days,
        })
        r += 4
    return people


def find_header_columns(ws, header_row=1):
    """Map known header labels (row 1) to their column index."""
    wanted = {
        "personal total": "total_meals",
        "personal cost(tk)": "personal_cost",
        "deposit(tk)": "deposit",
        "balance(tk)": "balance",
        "khalar bill": "khalar",
        "gas bill": "gas",
        "electricity bill": "electricity",
        "wifi bill": "wifi",
    }
    cols = {}
    for cell in ws[header_row]:
        key = norm(cell.value)
        if key in wanted:
            cols[wanted[key]] = cell.column
    return cols


def find_bazar_section(ws):
    """Locate the 'Bazar List' title, then its header row, then read rows
    until the Date column goes empty."""
    title_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) == "bazar list":
                title_cell = cell
                break
        if title_cell:
            break
    if not title_cell:
        return [], None

    header_row = title_cell.row + 1
    col_map = {}
    for cell in ws[header_row]:
        key = norm(cell.value)
        if key in ("date", "cost", "person", "main items"):
            col_map[key] = cell.column

    bazar = []
    r = header_row + 1
    blank_streak = 0
    while blank_streak < 3 and r < ws.max_row:
        date_val = ws.cell(row=r, column=col_map.get("date", 4)).value
        if date_val is None or str(date_val).strip() == "":
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val).strip()
        # normalize DD.MM.YYYY -> YYYY-MM-DD if needed
        if "." in date_str and len(date_str.split(".")) == 3:
            dd, mm, yyyy = date_str.split(".")
            date_str = f"{yyyy}-{mm}-{dd}"
        bazar.append({
            "date": date_str,
            "cost": to_float(ws.cell(row=r, column=col_map.get("cost", 9)).value),
            "person": str(ws.cell(row=r, column=col_map.get("person", 14)).value or "").strip(),
            "items": str(ws.cell(row=r, column=col_map.get("main items", 20)).value or "").strip(),
        })
        r += 1
    return bazar, title_cell.row


def main():
    xlsx_path = download_workbook(ONEDRIVE_SHARE_URL)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

    bazar, bazar_title_row = find_bazar_section(ws)
    bazar.sort(key=lambda b: b["date"])

    header_cols = find_header_columns(ws)
    people_raw = parse_people(ws, bazar_title_row or ws.max_row)

    people = []
    grand_total_meal = 0.0
    for p in people_raw:
        r = p["row"]
        total_meals = ws.cell(row=r, column=header_cols["total_meals"]).value if "total_meals" in header_cols else sum(d["total"] for d in p["days"].values())
        total_meals = to_float(total_meals, default=sum(d["total"] for d in p["days"].values()))
        grand_total_meal += total_meals

        def col_val(key, fallback=0.0):
            if key in header_cols:
                return to_float(ws.cell(row=r, column=header_cols[key]).value, fallback)
            return fallback

        people.append({
            "name": p["name"],
            "days": p["days"],
            "total_meals": total_meals,
            "personal_cost": round(col_val("personal_cost"), 2),
            "deposit": col_val("deposit"),
            "balance": round(col_val("balance"), 2),
            "khalar": col_val("khalar"),
            "gas": col_val("gas"),
            "electricity": col_val("electricity"),
        })

    meal_rate = to_float(find_label_value(ws, "Meal Rate"))
    total_bazar_cost = to_float(find_label_value(ws, "Total (Cost) Bazar")) or sum(b["cost"] for b in bazar)
    total_deposit = to_float(find_label_value(ws, "Total Deposit")) or sum(p["deposit"] for p in people)
    total_due = to_float(find_label_value(ws, "Total due"))
    current_balance = to_float(find_label_value(ws, "Current Balance"))
    grand_total_meal_label = to_float(find_label_value(ws, "Grand Total (Meal)"))
    if grand_total_meal_label:
        grand_total_meal = grand_total_meal_label

    data = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "meal_rate": round(meal_rate, 4),
        "grand_total_meal": grand_total_meal,
        "total_bazar_cost": round(total_bazar_cost, 2),
        "total_deposit": round(total_deposit, 2),
        "total_due": round(total_due, 2) if total_due else round(total_deposit - sum(p["personal_cost"] for p in people), 2),
        "current_balance": round(current_balance, 2),
        "people": sorted(people, key=lambda p: p["name"]),
        "bazar": bazar,
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Wrote data.json — {len(people)} people, {len(bazar)} bazar entries, "
          f"meal_rate={data['meal_rate']}, grand_total_meal={data['grand_total_meal']}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
