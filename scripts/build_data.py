#!/usr/bin/env python3
"""
Downloads the live workbook from OneDrive by following the share link redirects.
Works with standard "Anyone with the link" OneDrive shares.
"""
import base64
import json
import os
import sys
import tempfile
import urllib.request
import requests  # <-- we now use requests for better redirect handling
from datetime import datetime, timezone

import openpyxl

ONEDRIVE_SHARE_URL = os.environ["ONEDRIVE_SHARE_URL"]
SHEET_NAME = os.environ.get("SHEET_NAME")

DAY_COL_START = 2
DAY_COL_END = 32


def download_workbook(share_url: str) -> str:
    """Download the file from a public OneDrive share link."""
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
    
    # First, resolve the share link (it usually redirects to the real URL)
    resp = session.get(share_url, allow_redirects=False, timeout=30)
    
    # If it's a redirect (302/301), grab the Location header
    if resp.status_code in (301, 302):
        real_url = resp.headers.get("Location")
        if not real_url:
            raise Exception("Redirect URL not found in response headers")
        
        # Append &download=1 to force the browser to download the file
        if "?" in real_url:
            download_url = real_url + "&download=1"
        else:
            download_url = real_url + "?download=1"
        
        # Now download the actual file
        file_resp = session.get(download_url, timeout=60)
        file_resp.raise_for_status()
        
        tmp_path = os.path.join(tempfile.gettempdir(), "meal_chart_live.xlsx")
        with open(tmp_path, "wb") as f:
            f.write(file_resp.content)
        return tmp_path
    
    # Fallback: if no redirect, try the direct download trick (old method)
    b64 = base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    api_url = f"https://api.onedrive.com/v1.0/shares/u!{b64}/root/content"
    req = urllib.request.Request(api_url, headers={"User-Agent": "Mozilla/5.0"})
    tmp_path = os.path.join(tempfile.gettempdir(), "meal_chart_live.xlsx")
    with urllib.request.urlopen(req, timeout=60) as resp, open(tmp_path, "wb") as out:
        out.write(resp.read())
    return tmp_path


def to_float(v, default=0.0):
    try:
        if v is None or str(v).strip() in ("", "\u200b"):
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


def norm(s):
    return " ".join(str(s).split()).strip().lower() if s is not None else ""


def find_label_value(ws, label, max_cols_right=6, occurrence=0):
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) == norm(label):
                matches.append(cell)
    if not matches or occurrence >= len(matches):
        return None
    label_cell = matches[occurrence]
    for offset in range(1, max_cols_right + 1):
        v = ws.cell(row=label_cell.row, column=label_cell.column + offset).value
        if v is not None:
            return v
    return None


def parse_people(ws, bazar_start_row):
    people = []
    r = 2
    while r < bazar_start_row:
        name_cell = ws.cell(row=r, column=1).value
        name = str(name_cell).strip() if name_cell is not None else ""
        if not name or name.upper() == "N/A":
            r += 4
            continue

        def day_counts(row):
            return {
                str(c - DAY_COL_START + 1): to_float(ws.cell(row=row, column=c).value)
                for c in range(DAY_COL_START, DAY_COL_END + 1)
            }

        breakfast = day_counts(r)
        lunch = day_counts(r + 1)
        dinner = day_counts(r + 2)

        days = {}
        for d in breakfast.keys():
            b, l, din = breakfast[d], lunch.get(d, 0.0), dinner.get(d, 0.0)
            days[d] = {"breakfast": b, "lunch": l, "dinner": din, "total": b * 0.5 + l + din}

        people.append({"name": name, "row": r, "days": days})
        r += 4
    return people


def find_header_columns(ws, header_row=1):
    wanted = {
        "personal total": "total_meals",
        "personal cost(tk)": "personal_cost",
        "deposit(tk)": "deposit",
        "balance(tk)": "balance",
        "khalar bill": "khalar",
        "gas bill": "gas",
        "electricity bill": "electricity",
        "wifi bill": "wifi",
    }
    cols = {}
    for cell in ws[header_row]:
        key = norm(cell.value)
        if key in wanted:
            cols[wanted[key]] = cell.column
    return cols


def find_bazar_section(ws):
    title_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) == "bazar list":
                title_cell = cell
                break
        if title_cell:
            break
    if not title_cell:
        return [], None

    header_row = title_cell.row + 1
    col_map = {}
    for cell in ws[header_row]:
        key = norm(cell.value)
        if key in ("date", "cost", "person", "main items"):
            col_map[key] = cell.column

    bazar = []
    r = header_row + 1
    blank_streak = 0
    while blank_streak < 3 and r < ws.max_row:
        date_val = ws.cell(row=r, column=col_map.get("date", 4)).value
        if date_val is None or str(date_val).strip() == "":
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val).strip()
        if "." in date_str and len(date_str.split(".")) == 3:
            dd, mm, yyyy = date_str.split(".")
            date_str = f"{yyyy}-{mm}-{dd}"
        bazar.append({
            "date": date_str,
            "cost": to_float(ws.cell(row=r, column=col_map.get("cost", 9)).value),
            "person": str(ws.cell(row=r, column=col_map.get("person", 14)).value or "").strip(),
            "items": str(ws.cell(row=r, column=col_map.get("main items", 20)).value or "").strip(),
        })
        r += 1
    return bazar, title_cell.row


def main():
    xlsx_path = download_workbook(ONEDRIVE_SHARE_URL)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

    bazar, bazar_title_row = find_bazar_section(ws)
    bazar.sort(key=lambda b: b["date"])

    header_cols = find_header_columns(ws)
    people_raw = parse_people(ws, bazar_title_row or ws.max_row)

    people = []
    grand_total_meal = 0.0
    for p in people_raw:
        r = p["row"]
        total_meals = ws.cell(row=r, column=header_cols["total_meals"]).value if "total_meals" in header_cols else sum(d["total"] for d in p["days"].values())
        total_meals = to_float(total_meals, default=sum(d["total"] for d in p["days"].values()))
        grand_total_meal += total_meals

        def col_val(key, fallback=0.0):
            if key in header_cols:
                return to_float(ws.cell(row=r, column=header_cols[key]).value, fallback)
            return fallback

        people.append({
            "name": p["name"],
            "days": p["days"],
            "total_meals": total_meals,
            "personal_cost": round(col_val("personal_cost"), 2),
            "deposit": col_val("deposit"),
            "balance": round(col_val("balance"), 2),
            "khalar": col_val("khalar"),
            "gas": col_val("gas"),
            "electricity": col_val("electricity"),
            "wifi": col_val("wifi"),
        })

    meal_rate = to_float(find_label_value(ws, "Meal Rate"))
    total_bazar_cost = to_float(find_label_value(ws, "Total (Cost) Bazar")) or sum(b["cost"] for b in bazar)
    total_deposit = to_float(find_label_value(ws, "Total Deposit")) or sum(p["deposit"] for p in people)
    total_due = to_float(find_label_value(ws, "Total due"))
    current_balance = to_float(find_label_value(ws, "Current Balance"))
    grand_total_meal_label = to_float(find_label_value(ws, "Grand Total (Meal)"))
    if grand_total_meal_label:
        grand_total_meal = grand_total_meal_label

    data = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "meal_rate": round(meal_rate, 4),
        "grand_total_meal": grand_total_meal,
        "total_bazar_cost": round(total_bazar_cost, 2),
        "total_deposit": round(total_deposit, 2),
        "total_due": round(total_due, 2) if total_due else round(total_deposit - sum(p["personal_cost"] for p in people), 2),
        "current_balance": round(current_balance, 2),
        "people": sorted(people, key=lambda p: p["name"]),
        "bazar": bazar,
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Wrote data.json — {len(people)} people, {len(bazar)} bazar entries, "
          f"meal_rate={data['meal_rate']}, grand_total_meal={data['grand_total_meal']}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
